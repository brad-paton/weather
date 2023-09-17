
import requests
import openpyxl
from openpyxl.styles import Alignment



def clear_xlsx_file(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Iterate over all sheets in the workbook
    for sheet_name in workbook.sheetnames:
        # Select the sheet
        sheet = workbook[sheet_name]

        # Clear the contents of all cells in the sheet
        sheet.delete_rows(1, sheet.max_row)

    # Save the modified workbook
    workbook.save(file_path)


# Define your API key and the desired location
api_key = 'd8803f200694590d01f64d045af4efcf'
location = 'Orlando' # Replace with your desired location
path = 'C:\\Users\\bpato\\OneDrive\\weather\\Weathertest.xlsx'

clear_xlsx_file(path)

# Load the existing workbook
workbook = openpyxl.load_workbook(path)

# Select the active sheet
sheet = workbook.active

# Find the last row in the sheet
last_row = sheet.max_row + 1

# Make the API request to fetch the current weather data
current_weather_url = f'http://api.openweathermap.org/data/2.5/weather?q={location}&appid={api_key}&units=imperial'
response = requests.get(current_weather_url)
current_weather_data = response.json()

# Write the current weather data to the sheet
sheet.cell(row=last_row, column=1, value=current_weather_data['name'])
sheet.cell(row=last_row, column=2, value=current_weather_data['main']['temp'])
sheet.cell(row=last_row, column=3, value=current_weather_data['main']['humidity'])
sheet.cell(row=last_row, column=4, value=current_weather_data['weather'][0]['description'])

# Make the API request to fetch the forecasted weather data
forecast_url = f'http://api.openweathermap.org/data/2.5/forecast?q={location}&appid={api_key}&units=imperial'
response = requests.get(forecast_url)
forecast_data = response.json()

# Write the forecasted weather data to the sheet
row = last_row + 1
for forecast in forecast_data['list']:
    row += 1
    date = forecast['dt_txt']
    temp = forecast['main']['temp']
    humidity = forecast['main']['humidity']
    description = forecast['weather'][0]['description']

    sheet.cell(row=row, column=1, value=date)
    sheet.cell(row=row, column=2, value=temp)
    sheet.cell(row=row, column=3, value=humidity)
    sheet.cell(row=row, column=4, value=description)
    

# Save the updated workbook
workbook.save(path)
